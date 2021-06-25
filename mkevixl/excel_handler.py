from openpyxl import Workbook
import os

"""Excelに対する処理を行う"""


def make_new_workbook_name(filename):
    """Excelファイル名を返す

    Args:
        filename ([type]): [description]
    """

    if not filename.endswith(".xlsx"):
        # 拡張子がxlsx以外だったら強制的に変更する
        filename = os.path.splitext(filename)[0] + ".xlsx"

    # wb = Workbook()
    # wb.save(filename)

    return filename


class ExcelSheetByName:
    """指定した名称でシートを追加してExcelファイル保存"""

    def __init__(self, filename, sheetnames):
        """[summary]

        Args:
        filename ([string]): [excelfilename]
        sheetnames ([list]): [sheetname]
        """
        self.filename = filename
        self.sheetnames = sheetnames

    def execute(self):

        wb = Workbook()
        for sheetname in self.sheetnames:
            wb.create_sheet(sheetname)
        wb.remove(wb["Sheet"])

        return wb.save(self.filename)


class ExcelSheetByCount:
    """[指定した数字名と数だけシートを追加してExcelファイル保存]"""

    def __init__(self, filename, sheetcount, start):
        """[summary]

        Args:
            filename ([string]): [excelfilename]
            sheetcount ([int]): [add sheet count]
            start ([int]): [startsheetname]
        """
        self.filename = filename
        self.sheetcount = sheetcount
        self.start = start

    def execute(self):

        sheetname = self.start
        wb = Workbook()
        for count in range(0, self.sheetcount):
            wb.create_sheet(str(sheetname))
            sheetname = sheetname + 1
        wb.remove(wb["Sheet"])

        return wb.save(self.filename)


def sheettype():
    pass


def make_file(mode, filename, sheetcount, sheetnames):
    if mode == 1:
        print("mode 1")
        return ExcelSheetByName(filename, sheetnames).execute()
    elif mode == 2:
        return ExcelSheetByCount(filename, sheetcount, sheetnames).execute()


if __name__ == "__main__":
    # sheetlist = ["aiueo", "12345"]
    # test = ExcelSheetByName("test.xlsx", sheetlist)
    # test.add_work_sheet()

    # test = ExcelSheetByCount("countbyfile.xlsx", 10, 23)
    # test.execute()
    sheet = 10
    make_file(1, "test.xlsx", 10, sheet)
